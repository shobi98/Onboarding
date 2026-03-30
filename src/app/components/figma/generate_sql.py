import os
import re
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))

input_file = os.path.join(script_dir, "schema_cleaned.xlsx")
output_file = os.path.join(script_dir, "update_schema.sql")

wb = load_workbook(input_file)
ws = wb.active

header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

link_col = None
schema_col = None
for idx, name in enumerate(header_row):
    name_str = str(name or "").strip()
    if "Post" in name_str and "link" in name_str.lower():
        link_col = idx + 1
    if "Schema Code" in name_str:
        schema_col = idx + 1

if not link_col:
    print("'Post link' column not found!")
    exit(1)
if not schema_col:
    print("'Schema Code' column not found!")
    exit(1)

print(f"Post link column: {link_col}")
print(f"Schema Code column: {schema_col}")


def extract_short_id(url):
    """Extract short_id from URL like https://community.prepai.io/post/8sZZNh/..."""
    match = re.search(r'/post/([^/]+)', str(url))
    return match.group(1) if match else None


def extract_json_from_schema(text):
    """Extract JSON content, stripping <script> tags if present"""
    script_match = re.search(
        r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        text,
        re.S,
    )
    if script_match:
        return script_match.group(1).strip()
    return text.strip()


def escape_sql(text):
    """Escape single quotes for SQL"""
    return text.replace("\\", "\\\\").replace("'", "\\'")


sql_statements = []
skipped = []

for row in range(2, ws.max_row + 1):
    url = ws.cell(row, link_col).value
    schema = ws.cell(row, schema_col).value

    if not url or not schema:
        continue

    short_id = extract_short_id(url)
    if not short_id:
        skipped.append(f"Row {row}: Could not extract short_id from: {url}")
        continue

    json_content = extract_json_from_schema(str(schema))
    json_escaped = escape_sql(json_content)

    sql = f"UPDATE post SET structured_data = '{json_escaped}' WHERE short_id = '{short_id}';"
    sql_statements.append(sql)

with open(output_file, "w", encoding="utf-8") as f:
    f.write("-- Auto-generated SQL to update structured_data in post table\n")
    f.write(f"-- Total updates: {len(sql_statements)}\n\n")
    f.write("START TRANSACTION;\n\n")
    for stmt in sql_statements:
        f.write(stmt + "\n")
    f.write("\n-- Review the changes before committing\n")
    f.write("COMMIT;\n")

print(f"\nDone!")
print(f"SQL statements generated: {len(sql_statements)}")
if skipped:
    print(f"\nSkipped rows:")
    for s in skipped:
        print(f"  {s}")
print(f"\nSQL file saved to: {output_file}")
print(f"\nRun this SQL in phpMyAdmin or MySQL CLI to update the database.")
