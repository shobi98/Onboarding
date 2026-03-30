import os
import re
from openpyxl import load_workbook, Workbook

script_dir = os.path.dirname(os.path.abspath(__file__))

input_file = os.path.join(script_dir, "Discussion Schema Code - PrepAI.xlsx")
output_file = os.path.join(script_dir, "schema_cleaned.xlsx")


def clean_schema(text):
    # Replace "" (double double-quotes) with " (single double-quote)
    # But don't touch already-single quotes
    text = re.sub(r'"{2,}', '"', text)
    return text


wb_in = load_workbook(input_file)
ws_in = wb_in.active

header_row = [ws_in.cell(1, col).value for col in range(1, ws_in.max_column + 1)]

schema_col = None
for idx, name in enumerate(header_row):
    if name and "Schema Code" in str(name):
        schema_col = idx + 1
        break

if schema_col is None:
    print("'Schema Code' column not found!")
    exit(1)

print(f"Found 'Schema Code' in column {schema_col}")
print(f"Total rows: {ws_in.max_row - 1}")

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Cleaned Schema"

# Copy all headers from input
for col in range(1, ws_in.max_column + 1):
    ws_out.cell(1, col, ws_in.cell(1, col).value)

cleaned_count = 0

for row in range(2, ws_in.max_row + 1):
    for col in range(1, ws_in.max_column + 1):
        cell_value = ws_in.cell(row, col).value

        if col == schema_col and cell_value:
            original = str(cell_value)
            cleaned = clean_schema(original)
            ws_out.cell(row, col, cleaned)

            if original != cleaned:
                cleaned_count += 1
        else:
            ws_out.cell(row, col, cell_value)

wb_out.save(output_file)

print(f"\nDone!")
print(f"Total rows processed: {ws_in.max_row - 1}")
print(f"Rows where extra quotes were fixed: {cleaned_count}")
print(f"Output saved to: {output_file}")
